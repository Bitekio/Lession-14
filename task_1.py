"""
hw
"""
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

if __name__ == "__main__":
    current_directory = os.path.dirname(os.path.abspath(__file__))

    file_names = ['1111.xlsx', '2222.xlsx', '3333.xlsx']

    dfs = []

    for file_name in file_names:
        file_path = os.path.join(current_directory, file_name)
        df = pd.read_excel(file_path)
        dfs.append(df)

    for i, df in enumerate(dfs, start=1):
        dfs[i-1] = df.sort_values(by='ID', ascending=False)

    wb = Workbook()
    ws = wb.active

    font = Font(name='Arial', size=12, bold=True)
    border = Border(left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin'))

    for i, df in enumerate(dfs, start=1):
        for col_num, value in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=value)
            cell.font = font
            cell.border = border

        for row_num, (_, row) in enumerate(df.iterrows(), 2):
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border

    wb.save(os.path.join(current_directory, 'combined_data.xlsx'))
